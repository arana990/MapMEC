# Este arquivo é parte do programa Map&MEC
# Map&MEC é um software livre; você pode redistribuí-lo e/ou
# modificá-lo sob os termos da Licença Pública Geral GNU como publicada
# pela Free Software Foundation; na versão 3 da Licença, ou
# (a seu critério) qualquer versão posterior.
#
# Este programa é distribuído na esperança de que possa ser útil,
# mas SEM NENHUMA GARANTIA; sem uma garantia implícita de ADEQUAÇÃO
# a qualquer MERCADO ou APLICAÇÃO EM PARTICULAR. Veja a
# Licença Pública Geral GNU para mais detalhes.
#
# Você deve ter recebido uma cópia da Licença Pública Geral GNU junto
# com este programa. Se não, veja <http://www.gnu.org/licenses/>.

import pandas as pd
from unidecode import unidecode
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font
import streamlit as st
import base64
import io
# Definição funções

def remove_repetidos(lista):
    l = []
    for i in lista:
        if i not in l:
            l.append(i)
    return l

def get_value_from_siga(df, docente, column):
    mask = df['nome'] == docente
    values = df.loc[mask, column].values
    return values[0] if values.size > 0 else None

def get_value_from_progepe(df, docente, column):
    mask = df['NOME SERVIDOR'] == docente
    values = df.loc[mask, column].values
    return values[0] if values.size > 0 else None

def convert_vinculo(value):
    if value == "ATIVO PERMANENTE":
        return "Estatutário"
    elif value == "CONT.PROF.SUBSTITUTO":
        return "Substituto"
    elif value == "CONTR.PROF.VISITANTE":
        return "Visitante"
    else:
        return "Aposentado"  # Retorne o valor original se não corresponder a nenhuma das condições

def convert_titulacao(value):
    if value == 'DOUTORADO(T)' or value == 'DOUTORADO':
        return "Doutorado"
    elif value == 'MESTRADO(T)' or value == 'MESTRADO' or value == 'MESTRE+RSC-III (LEI 12772/12 ART 18)(T)':
        return "Mestrado"
    elif value == 'ESPECIALIZACAO NIVEL SUPERIOR(T)' or value == 'POS-GRADUAÇÃO+RSC-II LEI 12772/12 ART 18(T)':
        return "Especialização"
    elif value == 'ENSINO SUPERIOR' or value == 'GRADUACAO (NIVEL SUPERIOR COMPLETO)(T)':
        return "Graduação"
    else:
        return "Aposentado"

def convert_regime(value):
    if value == 'Dedc exclus':
        return "DE"
    elif value == '40 h sem':
        return "40h"
    elif value == '20 h sem':
        return "20h"
    else:
        return "Aposentado"


def get_data(serv_ativos, docente):
    mask = serv_ativos['NOME SERVIDOR'] == docente
    data = serv_ativos.loc[mask, 'DIA OCOR INGR ÓRGÃO EV'].values
    if len(data) > 0:
        # Supondo que data[0] seja uma string representando a data.
        # Convertendo para datetime object
        date_obj = pd.to_datetime(data[0])

        # Formatando o objeto de data para 'dia/mês/ano'
        formatted_date = date_obj.strftime('%d/%m/%Y')
        return formatted_date
    else:
        return 'Aposentado'

def generate_disciplinas(docentes, docente):
    mask = docentes['nome'] == docente
    dis = docentes.loc[mask, 'nomed'].values
    cod = docentes.loc[mask, 'codigo'].values
    list = ['%s - %s' %(cod[i],dis[i]) for i in range(dis.size)]
    return ', '.join(remove_repetidos(list))


# Função auxiliar para formatar as células do Excel
def format_excel_cells(cell, col_title):
    if col_title in ["Vínculo empregatício modelo Inep (não preencha esta coluna)",
                     "Tempo lecionando no curso (meses)  (não preencha esta coluna)"]:
        cell.fill = PatternFill(start_color="c00000", end_color="c00000", fill_type="solid")
    else:
        cell.fill = PatternFill(start_color="5b9bd5", end_color="5b9bd5", fill_type="solid")
    cell.font = Font(color="000000", bold=True)

def get_download_link(file_path, file_name):
    with open(file_path, 'rb') as f:
        file_data = f.read()
    b64 = base64.b64encode(file_data).decode()
    href = f'<a href="data:application/octet-stream;base64,{b64}" download="{file_name}">Clique aqui para baixar o arquivo</a>'
    return href


def process_file(uploaded_file):
    # Obtenha a data atual
    now = datetime.now()
    year = now.year

    # Dados da PROGEPE sobre os docentes
    progepe = pd.read_excel("DADOS_PROGEPE/01_rel_servidores_ativos_docentes.xlsx", skiprows=2)
    progepe['NOME SERVIDOR'] = progepe['NOME SERVIDOR'].apply(lambda x: unidecode(str(x)))
    progepe['ESCOLARIDADE'] = progepe['ESCOLARIDADE'].str.strip()

    # Dados do SIGA sobre os docentes
    #siga = pd.read_csv("DADOS_MEC/docentes.csv", encoding='ISO-8859-1', sep=';')
    siga = pd.read_csv(uploaded_file, encoding='ISO-8859-1', sep=';')
    siga['nome'] = siga['nome'].apply(unidecode)
    siga_3anos = siga[siga['ano'] >= (year-3)]
    st.write(siga_3anos.head())  # Exibir as primeiras linhas apenas para demonstração
    docentes_3anos = remove_repetidos(siga_3anos['nome'].values)

    df = pd.DataFrame()
    df['NOME DO DOCENTE E/OU TUTOR DO CURSO (Nome completo)'] = docentes_3anos
    df['LOTAÇÃO'] = [get_value_from_siga(siga_3anos, docente, 'lotacao') for docente in docentes_3anos]
    df['CPF'] = [get_value_from_siga(siga_3anos, docente, 'documento') for docente in docentes_3anos]
    df['lattes'] = [get_value_from_siga(siga_3anos, docente, 'curriculoLattes') for docente in docentes_3anos]
    df['E-mail'] = [get_value_from_siga(siga_3anos, docente, 'email') for docente in docentes_3anos]
    df['TITULAÇÃO MÁXIMA'] = [convert_titulacao(get_value_from_progepe(progepe, docente, 'ESCOLARIDADE')) for docente in docentes_3anos]
    df['REGIME DE TRABALHO UFPR'] = [convert_regime(get_value_from_progepe(progepe, docente, 'JORNADA TRABALHO')) for docente in docentes_3anos]
    df['DISCIPLINAS QUE LECIONA/LECIONOU'] = [generate_disciplinas(siga_3anos, docente) for docente in docentes_3anos]


    df['VÍNCULO EMPREGATÍCIO COM A UFPR'] = [convert_vinculo(get_value_from_progepe(progepe, docente, 'SITUAÇÃO VÍNCULO')) for docente in docentes_3anos]
    conditions = [
        df['VÍNCULO EMPREGATÍCIO COM A UFPR'] == "Estatutário",
        df['VÍNCULO EMPREGATÍCIO COM A UFPR'] == "Outro",
        df['VÍNCULO EMPREGATÍCIO COM A UFPR'] == "Substituto",
        df['VÍNCULO EMPREGATÍCIO COM A UFPR'] == "Visitante"
    ]
    choices = ["Estatutário", "Outro", "Outro", "Outro"]

    df["Vínculo empregatício modelo Inep (não preencha esta coluna)"] = np.select(conditions, choices, default="Outro")

    df['INGRESSO COMO DOCENTE DO CURSO (data exata)'] = [get_data(progepe, docente) for docente in docentes_3anos]
    df['INGRESSO COMO DOCENTE DO CURSO (datetime)'] = pd.to_datetime(df['INGRESSO COMO DOCENTE DO CURSO (data exata)'], format='%d/%m/%Y', errors='coerce')
    df["Tempo lecionando no curso (meses)  (não preencha esta coluna)"] = (pd.Timestamp.now() - df['INGRESSO COMO DOCENTE DO CURSO (datetime)']).dt.days // 30

    # Lendo todas as folhas do arquivo uma vez
    mec_docentes1 = pd.read_excel("DADOS_MEC/Dados_docentes_MEC.xlsx", sheet_name='Dados docentes', skiprows=1)
    cols_to_drop = [col for col in mec_docentes1.columns if 'Unnamed' in col]
    mec_docentes1 = mec_docentes1.drop(columns=cols_to_drop)

    mec_docentes2 = pd.read_excel("DADOS_MEC/Dados_docentes_MEC.xlsx", sheet_name='Dados docentes', skiprows=2)
    cols_to_drop = [col for col in mec_docentes2.columns if 'Unnamed' in col]
    mec_docentes2 = mec_docentes2.drop(columns=cols_to_drop)

    merged_dataframe = pd.concat([mec_docentes1, mec_docentes2], ignore_index=True)
    df = df.reindex(columns=merged_dataframe.columns).fillna('')

    # Criando um novo workbook e uma nova worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = 'Dados docentes'

    # Dicionário para armazenar a largura máxima de cada coluna
    max_col_widths = {}

    # Preenchendo a worksheet com os dados do DataFrame
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell_width = len(str(cell.value))
            col_title = ws.cell(row=1, column=c_idx).value

            if r_idx == 1:  # Se for o cabeçalho (linha de título)
                format_excel_cells(cell, col_title)

            # Atualizando a largura máxima da coluna, se necessário
            max_col_widths[cell.column_letter] = max(max_col_widths.get(cell.column_letter, 0), cell_width)

    # Ajustando a largura de cada coluna
    for col_letter, col_width in max_col_widths.items():
        ws.column_dimensions[col_letter].width = col_width + 2  # Adicione 2 para dar um pouco de espaço extra

    output = io.BytesIO()
    wb.save(output)
    return output

# Função para mostrar o tutorial
def show_tutorial():
    st.title("Tutorial: Como obter dados do SIGA")

    st.write("Siga os passos abaixo para aprender como extrair os dados dos docentes diretamente do SIGA:")

    # Carregando e mostrando a primeira imagem
    st.subheader("Passo 1: ")
    st.write("Clique em \"Relatórios\" e procure a opção \"Docentes Coordenação\" ")
    image1 = st.image("IMAGEM/parte01.png", use_column_width=True)

    # Carregando e mostrando a segunda imagem
    st.subheader("Passo 2")
    st.write("Selecione ⚠️TODAS⚠️ as colunas disponíveis. Em seguida selecione a opção Exportar CSV.")
    image2 = st.image("IMAGEM/parte02.png", use_column_width=True)

    st.write("Após seguir todos os passos, você deverá ter um arquivo CSV pronto para ser carregado na aplicação!")
